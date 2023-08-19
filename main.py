import openpyxl
from googletrans import Translator
from textblob import Word
import requests

def translate():
    colon_to_translate = str(input("colon to translate: "))
    if len(colon_to_translate)!=1:
        print("wrong symbol!!!")
        translate()
    colon_with_text = str(input("colon with text: "))
    if len(colon_with_text)!=1:
        print("wrong symbol!!!")
        translate()

    translator = Translator()
    for i in range(worksheet.max_row):
        if worksheet[colon_to_translate][i].value is None:
            if worksheet[colon_with_text][i].value is not None:
                if type(worksheet[colon_with_text][i].value) is str:
                    if len(worksheet[colon_with_text][i].value) < 25:
                        print(str(colon_with_text) + ": " + str(worksheet["A"][i].value))
                        print(str(colon_to_translate) + ": " + str(worksheet["L"][i].value) + "\n")
                        translation_ru = translator.translate(worksheet[colon_with_text][i].value, dest="ru")
                        translation_eng = translator.translate(worksheet[colon_with_text][i].value, dest="en")

                        full_translation = str(translation_eng.text) + "(" + str(translation_ru.pronunciation) + ")"
                        print(full_translation + "\n\n")

                        worksheet[colon_to_translate][i].value = full_translation

                    else:
                        print("\n\tERROR: the " + str(colon_with_text) + " line, row: " + str(i) + "is more than 25 symbols\n")
                else:
                    print("\n\tERROR: the " + str(colon_with_text) + " line, row: " + str(i) + "is not a str value\n")

    try:
        wb.save(FILE_PATH + FILE_NAME)
        print("File is saved!!!")
    except PermissionError:
        print('File might be opened, please close it before writing')

def print_file():
    for row in worksheet.iter_rows():
        text_row = str(row) + '.\t'
        for cell in row:
            text_row = str(text_row) + str(cell.value) + '\t'

        print(text_row)


FILE_PATH = 'C:/Users/6WIRI/Downloads/'
FILE_NAME = 'Платки_2022.xlsx'

wb = openpyxl.load_workbook(FILE_PATH + FILE_NAME)

print(wb.sheetnames)
inp = str(input('sheetname:'))

worksheet = wb[inp]


while True:
    inp = int(input("what would you like to do \n\t1. translate\n"))

    if (inp == 1):
        translate()
    elif (inp == 2):
        print_file()
    else:
        break



